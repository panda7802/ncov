# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import logging
import os

import openpyxl as openpyxl
import scrapy
from openpyxl import Workbook


class HisItem(scrapy.Item):
    # 时间
    time_now = scrapy.Field()
    # 地区
    province = scrapy.Field()
    # 确诊
    sure = scrapy.Field()
    # 死亡
    dead = scrapy.Field()
    # 治愈
    ok = scrapy.Field()
    # 疑似
    maybe = scrapy.Field()


class Ncov2019Pipeline(object):

    def __init__(self):
        logging.debug("-------HisPipeline __init__---------")
        self.file_full_dir = "./res_dir"
        self.file_full_path = self.file_full_dir + "/ncov.xlsx"
        self.types = []
        if not os.path.exists(self.file_full_path):
            self.wb = Workbook()  # class实例化
            self.ws = self.wb.active  # 激活工作表
            if not os.path.exists(self.file_full_dir):
                os.makedirs(self.file_full_dir)
            self.wb.save(self.file_full_path)  # 保存文件
        else:
            self.wb = openpyxl.load_workbook(self.file_full_path)
            self.ws = self.wb.active  # 激活工作表

    def init_ws(self, province_name):
        # type_index = self.types.index(type_name)
        # self.ws = self.wb.active
        self.ws = self.wb.create_sheet(province_name)
        self.ws.title = province_name
        self.ws.append(['时间', '范围', '确诊', '死亡', '治愈', '疑似'])  # 加入一行数据

    def open_spider(self, spider):
        logging.debug("-------HisPipeline open_spider---------")

    def add_line(self, item):
        """
        增加行
        :param item:
        :return:
        """
        line = [item['time_now'], item['province'], int(item['sure']), int(item['dead']), int(item['ok']),
                int(item['maybe'])]
        province_name = item['province']
        try:
            self.ws = self.wb.get_sheet_by_name(province_name)
        except Exception:
            self.init_ws(province_name)
            self.ws = self.wb.get_sheet_by_name(province_name)
        self.ws.append(line)
        # name_cell = "A%d" % self.ws.max_row
        # self.ws[name_cell].hyperlink = item['url']
        self.wb.save(self.file_full_path)  # 保存文件

    def process_item(self, item, spider):
        self.add_line(item)
        return item

    def close_spider(self, spider):
        self.wb.save(self.file_full_path)
